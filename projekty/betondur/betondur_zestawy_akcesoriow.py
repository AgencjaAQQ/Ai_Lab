import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─── kolory ───────────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # granat
C_HEADER_FONT = "FFFFFF"
C_CAT_BG      = "2E74B5"   # niebieski kategoria
C_CAT_FONT    = "FFFFFF"
C_PROD_BG     = "D6E4F0"   # jasnoniebieski – wiersz produktu
C_ACC_BG      = "EBF3FB"   # bardzo jasny – wiersz akcesorium
C_UZAS_BG     = "FFF2CC"   # żółtawy – uzasadnienie
C_ALT_ACC_BG  = "F2F7FC"   # alternating row
C_BORDER      = "B0C4D8"

thin = Side(style='thin', color=C_BORDER)
full_border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_fill(color): return PatternFill("solid", fgColor=color)
def hdr_font(color, bold=True, size=10): return Font(color=color, bold=bold, size=size)
def center(wrap=True): return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
def left(wrap=True):   return Alignment(horizontal='left',   vertical='center', wrap_text=wrap)

# ─────────────────────────────────────────────────────────────────────────────
# ARKUSZ 1 – ZESTAWY AKCESORIÓW
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Zestawy akcesoriów"
ws.freeze_panes = "A3"

# nagłówki kolumn
columns = [
    ("Kategoria produktu Betondur",  28),
    ("Produkt główny",               34),
    ("Akcesorium – propozycja",      36),
    ("Cena akcesorium",              16),
    ("URL akcesorium",               52),
    ("Priorytet\n(1=obowiązkowe)", 14),
    ("Uzasadnienie",                 58),
]

# wiersz 1 – tytuł
ws.merge_cells("A1:G1")
title_cell = ws["A1"]
title_cell.value = "BETONDUR.EU – Propozycje zestawów akcesoriów do kategorii produktów"
title_cell.fill  = hdr_fill(C_HEADER_BG)
title_cell.font  = Font(color=C_HEADER_FONT, bold=True, size=13)
title_cell.alignment = center()
ws.row_dimensions[1].height = 30

# wiersz 2 – nagłówki
for col_idx, (col_name, col_width) in enumerate(columns, start=1):
    cell = ws.cell(row=2, column=col_idx, value=col_name)
    cell.fill      = hdr_fill(C_CAT_BG)
    cell.font      = hdr_font(C_HEADER_FONT)
    cell.alignment = center()
    cell.border    = full_border
    ws.column_dimensions[get_column_letter(col_idx)].width = col_width
ws.row_dimensions[2].height = 38

# ─── dane ─────────────────────────────────────────────────────────────────────
# Struktura: (kategoria, produkt_glowny, [ (akcesorium, cena, url, priorytet, uzasadnienie) ])
data = [

    # ══════════════════════════════════════════════════════════════════════════
    # 1. BETON efekt dekoracyjny
    # ══════════════════════════════════════════════════════════════════════════
    (
        "BETON\nEfekt dekoracyjny",
        "Betondur BETON\nefekt dekoracyjny\ndo ścian",
        [
            ("Hardy Paca Stucco (wenecka) seria 35\n24×9 cm uchwyt 2K",
             "34,51 zł",
             "https://betondur.eu/produkt/hardy-paca-stucco-wenecka-seria-35-24x9-cm-uchwyt-2k/",
             "1 – obowiązkowe",
             "Masa BETON nakładana jest wyłącznie pacą wenecką warstwą ~1 mm. Bez odpowiedniej pacy niemożliwe jest uzyskanie charakterystycznego efektu betonu. Seria 35 (24×9 cm) to optymalny rozmiar do ścian – większa powierzchnia = szybsza i bardziej równomierna aplikacja."),
            ("Hardy Wałek malarski 25 cm\nruno 12 mm Ø48 mm",
             "20,06 zł",
             "https://betondur.eu/produkt/hardy-walek-malarski-25-cm-runo-12-mm-o48-mm/",
             "1 – obowiązkowe",
             "Podkład pod BETON (szary) i lakier zabezpieczający nakłada się wałkiem. Runo 12 mm zapewnia równomierne krycie na ścianach bez nadmiernego chłonięcia materiału. Wałek 25 cm skraca czas pracy na dużych powierzchniach."),
            ("Scley Taśma malarska precyzyjna\nWashi 572, 48 mm × 33 m",
             "14,78 zł",
             "https://betondur.eu/produkt/scley-tasma-malarska-precyzyjna-washi-572-48-mm-x-33-m-do-60-dni/",
             "1 – obowiązkowe",
             "Przy efektach dekoracyjnych precyzyjne maskowanie narożników, listew i okien jest krytyczne. Taśma Washi (do 60 dni) nie niszczy podłoża i daje idealnie proste linie. Szerokość 48 mm = lepsze zabezpieczenie przy nakładaniu masy pacą."),
            ("Hardy Folia ochronna HDPE\n4×5 m Standard",
             "15,97 zł",
             "https://betondur.eu/produkt/hardy-folia-ochronna-hdpe-4x5-m-standard/",
             "1 – obowiązkowe",
             "Masa BETON jest trudna do usunięcia z podłóg i mebli. Folia HDPE (mocniejsza, grubość standard) chroni całe pomieszczenie podczas nanoszenia masy i lakieru. Klienci dekoracyjni to często osoby prywatne bez doświadczenia – folia zabezpiecza ich przed kosztownymi uszkodzeniami."),
            ("Hardy Kuweta malarska\n26×35 cm (4\")",
             "7,44 zł",
             "https://betondur.eu/produkt/hardy-kuweta-malarska-26x35-cm-4%e2%80%b3/",
             "2 – zalecane",
             "Do nakładania lakieru i podkładu wałkiem niezbędna jest kuweta. Wkładka do kuwety (3,03 zł) ułatwia czyszczenie – warto oferować jako parę. Mała cena = wysoka konwersja dosprzedaży."),
            ("Hardy Pędzel płaski M7\n40 mm seria 37",
             "5,51 zł",
             "https://betondur.eu/produkt/hardy-pedzel-plaski-m7-40-mm-seria-37/",
             "2 – zalecane",
             "Narożniki ścian, miejsca przy listwach i sufitem wymagają precyzyjnego nanoszenia podkładu i lakieru pędzlem. Produkt do efektów dekoracyjnych = praca detalistyczna. Niska cena podnosi AOV bez oporu klienta."),
        ]
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # 2. RDZA efekt dekoracyjny
    # ══════════════════════════════════════════════════════════════════════════
    (
        "RDZA\nEfekt dekoracyjny",
        "Betondur RDZA\nefekt dekoracyjny\ndo ścian",
        [
            ("Hardy Paca Stucco (wenecka) seria 25\n20×8 cm uchwyt drewniany",
             "29,13 zł",
             "https://betondur.eu/produkt/hardy-paca-stucco-wenecka-seria-25-20x8-cm-uchwyt-drewniany/",
             "1 – obowiązkowe",
             "RDZA nakładana jest techniką pacą wenecką nieregularnymi ruchami – to fundamentalne narzędzie. Seria 25 (mniejsza, 20×8 cm) jest idealna do efektu rdzy, bo nieregularne, chaotyczne ruchy dają bardziej naturalny efekt. Bez pacy weneckiej produktu nie można zastosować."),
            ("Hardy Gąbka glazurnicza\n23×12×7 cm",
             "24,49 zł",
             "https://betondur.eu/produkt/hardy-gabka-glazurnicza-23x12x7-cm/",
             "1 – obowiązkowe",
             "Aktywator do efektu rdzy spryskuje się lub rozprowadza gąbką w nieregularny sposób – to właśnie tą techniką uzyskuje się autentyczny efekt korozji. Gąbka glazurnicza (gruba, strukturalna) jest standardowym narzędziem przy tej technice. Bez niej efekt wygląda sztucznie."),
            ("Scley Taśma malarska precyzyjna\nWashi 572, 30 mm × 33 m",
             "9,07 zł",
             "https://betondur.eu/produkt/scley-tasma-malarska-precyzyjna-washi-572-30-mm-x-33-m-do-60-dni/",
             "1 – obowiązkowe",
             "RDZA to efekt w stylu loft/industrial – często nakładany na wybrany fragment ściany jako accent wall. Precyzyjne granice efektu wymagają profesjonalnej taśmy Washi. Uniknięcie plam rdzy na sąsiednich powierzchniach jest krytyczne (aktywator jest agresywny chemicznie)."),
            ("Hardy Folia ochronna LDPE\n4×5 m z regranulatu",
             "8,63 zł",
             "https://betondur.eu/produkt/hardy-folia-ochronna-ldpe-4x5-m-z-regranulatu/",
             "1 – obowiązkowe",
             "Aktywator do efektu rdzy może trwale zabarwić podłogę lub meble. Folia ochronna jest absolutnym must-have przy tej linii. LDPE (tańsza) wystarczy do ochrony podłogi podczas nakładania masy."),
            ("Hardy Pędzel płaski M7\n40 mm seria 37",
             "5,51 zł",
             "https://betondur.eu/produkt/hardy-pedzel-plaski-m7-40-mm-seria-37/",
             "2 – zalecane",
             "Podkład i lakier w narożnikach nakładany pędzlem. Przy linii RDZA warto też użyć pędzla do rozprofilowania efektu przy krawędziach. Bardzo niska cena = łatwy cross-sell."),
            ("Hardy Kuweta malarska\n26×35 cm (4\")",
             "7,44 zł",
             "https://betondur.eu/produkt/hardy-kuweta-malarska-26x35-cm-4%e2%80%b3/",
             "2 – zalecane",
             "Do nakładania podkładu i lakieru zabezpieczającego wałkiem. Uzupełnia zestaw – bez kuwety wałek jest bezużyteczny."),
        ]
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # 3. OBSYDIAN efekt dekoracyjny
    # ══════════════════════════════════════════════════════════════════════════
    (
        "OBSYDIAN\nEfekt dekoracyjny",
        "Betondur OBSYDIAN\nefekt dekoracyjny\ndo ścian (czarny)",
        [
            ("Hardy Paca Stucco (wenecka) seria 35\n24×9 cm uchwyt 2K",
             "34,51 zł",
             "https://betondur.eu/produkt/hardy-paca-stucco-wenecka-seria-35-24x9-cm-uchwyt-2k/",
             "1 – obowiązkowe",
             "OBSYDIAN to masa strukturalna z ziarnistością ~1 mm, nakładana pacą wenecką. Seria 35 (największa z oferty) to optymalny wybór – większa paca = efektywniejsze tworzenie charakterystycznych wgłębień w strukturze kamienia wulkanicznego. Bez pacy produkt jest nienakladalny."),
            ("Hardy Wałek malarski Hardstar\n25 cm runo 13 mm Ø48 mm",
             "17,81 zł",
             "https://betondur.eu/produkt/hardy-walek-malarski-hardstar-25-cm-runo-13-mm-o48-mm/",
             "1 – obowiązkowe",
             "Podkład pod OBSYDIAN nakłada się wałkiem. Runo 13 mm (grubsze niż standardowe) lepiej sprawdza się przy podkładach akrylowych, zapewniając równomierne krycie. Marka Hardstar = lepsza trwałość włókien przy produktach chemicznych."),
            ("Scley Taśma malarska precyzyjna\nWashi 572, 48 mm × 33 m",
             "14,78 zł",
             "https://betondur.eu/produkt/scley-tasma-malarska-precyzyjna-washi-572-48-mm-x-33-m-do-60-dni/",
             "1 – obowiązkowe",
             "OBSYDIAN to efekt premium (czarny, błyszczący) – kupują go klienci dbający o detal. Precyzyjne linie maskowania są dla nich absolutnym priorytetem. Taśma szeroka 48 mm chroni więcej powierzchni przy nakładaniu grubej, strukturalnej masy."),
            ("Hardy Folia ochronna HDPE\n4×5 m Standard",
             "15,97 zł",
             "https://betondur.eu/produkt/hardy-folia-ochronna-hdpe-4x5-m-standard/",
             "1 – obowiązkowe",
             "Masa OBSYDIAN (ciemna, strukturalna) bardzo trudna do usunięcia z powierzchni – HDPE (mocniejsza folia) jest tu konieczna. Klienci premium docenią profesjonalne podejście do ochrony przestrzeni."),
            ("Hardy Pędzel płaski M7\n40 mm seria 37",
             "5,51 zł",
             "https://betondur.eu/produkt/hardy-pedzel-plaski-m7-40-mm-seria-37/",
             "2 – zalecane",
             "Do aplikacji podkładu i lakieru przy krawędziach i narożnikach. Precyzja wykończenia jest szczególnie ważna przy czarnym efekcie – każda nierównomierna krawędź będzie widoczna."),
            ("Hardy Kuweta ręczna malarska\n1 l do wałka 15 cm",
             "27,83 zł",
             "https://betondur.eu/produkt/hardy-kuweta-reczna-malarska-1-l-do-walka-15-cm/",
             "2 – zalecane",
             "Kuweta ręczna (z rączką) jest wygodniejsza do pracy z lakierem i podkładem OBSYDIAN, ponieważ często pracuje się w trudno dostępnych miejscach. Model premium dopasowany do klienta kupującego produkt z wyższej półki."),
        ]
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # 4. HARDFLOOR – farba do posadzek
    # ══════════════════════════════════════════════════════════════════════════
    (
        "HARDFLOOR\nFarba do posadzek",
        "Betondur HARDFLOOR\nfarba do posadzek\ni betonu",
        [
            ("Hardy Zestaw malarski nr 64\n(wałek Hardstar 25 cm, runo 13 mm)",
             "47,45 zł",
             "https://betondur.eu/produkt/hardy-zestaw-malarski-nr-64-walek-hardstar-25-cm-runo-13-mm/",
             "1 – obowiązkowe",
             "Farba do posadzek nakłada się wałkiem na duże powierzchnie poziome. Zestaw malarski nr 64 (wałek + wkładki + kij) to kompletne narzędzie – klient nie musi szukać pasujących komponentów osobno. Runo 13 mm jest optymalne do gęstych farb posadzkowych. Garaże i magazyny = duże powierzchnie = wałek 25 cm obowiązkowy."),
            ("Scley Taśma malarska precyzyjna\nWashi 572, 48 mm × 33 m",
             "14,78 zł",
             "https://betondur.eu/produkt/scley-tasma-malarska-precyzyjna-washi-572-48-mm-x-33-m-do-60-dni/",
             "1 – obowiązkowe",
             "Przy malowaniu posadzek konieczne jest precyzyjne maskowanie krawędzi przy ścianach, progach, słupkach. Szeroka taśma 48 mm zapewnia lepszą ochronę przy malowaniu wałkiem. Farba do garażu musi mieć równe krawędzie – szczególnie w pobliżu ścian i bram."),
            ("Hardy Pędzel płaski M7\n40 mm seria 37",
             "5,51 zł",
             "https://betondur.eu/produkt/hardy-pedzel-plaski-m7-40-mm-seria-37/",
             "1 – obowiązkowe",
             "Narożniki podłogi przy ścianach (podbitka) malowane są pędzlem, nie wałkiem. Bez pędzla klient nie wykończy narożników profesjonalnie. Niska cena = oczywisty cross-sell, który realnie ułatwia pracę."),
            ("Hardy Folia ochronna LDPE\n4×5 m z regranulatu",
             "8,63 zł",
             "https://betondur.eu/produkt/hardy-folia-ochronna-ldpe-4x5-m-z-regranulatu/",
             "2 – zalecane",
             "Ochrona ścian i mebli przy malowaniu posadzki. Szczególnie ważne gdy farba posadzkowa nakładana jest w garażu lub piwnicy z pozostawionymi przedmiotami. Niska cena = łatwa decyzja zakupowa."),
            ("Hardy Kratka malarska\ndo wiadra 27×32 cm",
             "4,53 zł",
             "https://betondur.eu/produkt/hardy-kratka-malarska-do-wiadra-27x32-cm/",
             "2 – zalecane",
             "Kratka do wiadra jest praktyczniejsza od kuwety przy malowaniu posadzek – wiadro można przenosić po podłodze, a kratka odciska nadmiar farby z wałka. Idealne przy malowaniu garażu, gdzie kubatura jest duża i kuweta byłaby nieporęczna."),
        ]
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # 5. EKO – farba do betonu (ściany)
    # ══════════════════════════════════════════════════════════════════════════
    (
        "EKO\nFarba do betonu",
        "Betondur EKO\nfarba do betonu\n(ściany wewn./zewn.)",
        [
            ("Hardy Zestaw malarski nr 64\n(wałek Hardstar 25 cm, runo 13 mm)",
             "47,45 zł",
             "https://betondur.eu/produkt/hardy-zestaw-malarski-nr-64-walek-hardstar-25-cm-runo-13-mm/",
             "1 – obowiązkowe",
             "EKO to farba do piwnic, garaży, kotłowni – aplikowana wałkiem na dużych powierzchniach. Zestaw nr 64 (kompletny: trzonek + wałek + kuweta) to optymalny wybór – klient dostaje wszystko w jednym. Runo 13 mm dobrze sprawdza się przy chropowatych powierzchniach betonowych."),
            ("Hardy Pędzel płaski M7\n40 mm seria 37",
             "5,51 zł",
             "https://betondur.eu/produkt/hardy-pedzel-plaski-m7-40-mm-seria-37/",
             "1 – obowiązkowe",
             "Narożniki, łączenia ze stropem, słupy – wszystkie miejsca niedostępne dla wałka wymagają pędzla. Przy malowaniu piwnic i garaży pędzel to niezbędne uzupełnienie wałka. Bardzo niska cena = łatwy cross-sell."),
            ("Scley Taśma malarska precyzyjna\nWashi 572, 30 mm × 33 m",
             "9,07 zł",
             "https://betondur.eu/produkt/scley-tasma-malarska-precyzyjna-washi-572-30-mm-x-33-m-do-60-dni/",
             "2 – zalecane",
             "Maskowanie okien, rur, instalacji w piwnicach i garażach. Taśma 30 mm jest wystarczająca przy standardowych pracach malarskich (w odróżnieniu od efektów dekoracyjnych, gdzie potrzebna jest szersza). Tańsza wersja = pasuje do budżetu klientów kupujących praktyczną farbę techniczną."),
            ("Hardy Folia ochronna LDPE\n4×5 m z regranulatu",
             "8,63 zł",
             "https://betondur.eu/produkt/hardy-folia-ochronna-ldpe-4x5-m-z-regranulatu/",
             "2 – zalecane",
             "Ochrona podłóg, urządzeń i instalacji w piwnicach/garażach/kotłowniach podczas malowania. Tańsza folia LDPE wystarczy do tego zastosowania – dopasowana cenowo do produktu technicznego, nie premium."),
        ]
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # 6. OUTDOOR – powłoka zewnętrzna
    # ══════════════════════════════════════════════════════════════════════════
    (
        "OUTDOOR\nPowłoka zewnętrzna",
        "Betondur OUTDOOR\npowłoka akrylowa\nna beton (zewnątrz)",
        [
            ("Hardy Wałek malarski 25 cm\nruno 12 mm Ø48 mm",
             "20,06 zł",
             "https://betondur.eu/produkt/hardy-walek-malarski-25-cm-runo-12-mm-o48-mm/",
             "1 – obowiązkowe",
             "OUTDOOR to powłoka do tarasów, balkonów, parkingów – duże powierzchnie poziome i pionowe. Wałek 25 cm to standardowe narzędzie do zewnętrznych nawierzchni. Runo 12 mm dobrze wchodzi w teksturę betonu zewnętrznego."),
            ("Hardy Zestaw malarski nr 68\n(wałek Hardmax 18 cm, runo 18 mm)",
             "30,20 zł",
             "https://betondur.eu/produkt/hardy-zestaw-malarski-nr-68-walek-hardmax-18-cm-runo-18-mm/",
             "1 – obowiązkowe",
             "Wałek z grubym runem 18 mm jest szczególnie skuteczny na chropowatych powierzchniach zewnętrznych (beton, asfalt). Zestaw kompletny z trzonkiem i kuwetą. Komplementarny z wałkiem 25 cm – razem tworzą optymalny zestaw do zewnętrznych nawierzchni."),
            ("Scley Taśma malarska precyzyjna\nWashi 572, 48 mm × 33 m",
             "14,78 zł",
             "https://betondur.eu/produkt/scley-tasma-malarska-precyzyjna-washi-572-48-mm-x-33-m-do-60-dni/",
             "1 – obowiązkowe",
             "Maskowanie obrzeży tarasu, krawędzi balkonu, elewacji przy malowaniu powłoką zewnętrzną. Taśma Washi (do 60 dni) zachowuje właściwości na zewnątrz, gdzie prace mogą być przerywane przez warunki pogodowe."),
            ("Hardy Pędzel płaski M7\n40 mm seria 37",
             "5,51 zł",
             "https://betondur.eu/produkt/hardy-pedzel-plaski-m7-40-mm-seria-37/",
             "2 – zalecane",
             "Wykończenie krawędzi, malowanie narożników, obrzeży i miejsc niedostępnych dla wałka na zewnątrz. Pędzel odporny na warunki zewnętrzne."),
            ("Scley Tektura malarska falista\n150 g/m², 1×15 m",
             "26,85 zł",
             "https://betondur.eu/produkt/scley-tektura-malarska-falista-150-g-m%c2%b2-1-x-15-m/",
             "2 – zalecane",
             "Przy malowaniu tarasów i balkonów konieczna jest ochrona elewacji, mebli ogrodowych, roślin w donicach. Tektura malarska (15 m!) zapewnia szeroką ochronę – bardziej trwała niż folia przy pracach zewnętrznych z wiatrem."),
        ]
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # 7. LINER – farba do znakowania
    # ══════════════════════════════════════════════════════════════════════════
    (
        "LINER\nFarba do znakowania",
        "Betondur LINER\nfarba do znakowania\npoziomego",
        [
            ("Scley Taśma malarska precyzyjna\nWashi 572, 48 mm × 33 m",
             "14,78 zł",
             "https://betondur.eu/produkt/scley-tasma-malarska-precyzyjna-washi-572-48-mm-x-33-m-do-60-dni/",
             "1 – obowiązkowe",
             "Linie parkingowe, oznaczenia pasów, linie ewakuacyjne – wszystkie wymagają idealnie prostych krawędzi. Taśma szeroka 48 mm jest wręcz stworzona do wyznaczania pasów przy malowaniu poziomym. To najważniejsze akcesorium do LINER – bez precyzyjnej taśmy linie będą krzywe i nieprofesjonalne."),
            ("Hardy Wałek malarski Hardstar\n10 cm runo 9 mm",
             "11,87 zł",
             "https://betondur.eu/produkt/hardy-walek-malarski-hardstar-10-cm-runo-9-mm/",
             "1 – obowiązkowe",
             "Mały wałek 10 cm jest idealny do malowania wąskich linii parkingowych i oznaczeń poziomych – bardziej precyzyjny niż wałek 25 cm, mieszczący się dokładnie między dwiema taśmami maskującymi. Runo 9 mm (cienkie) = równomierny naniesienie gęstej farby LINER na gładkich powierzchniach."),
            ("Hardy Pędzel płaski M7\n40 mm seria 37",
             "5,51 zł",
             "https://betondur.eu/produkt/hardy-pedzel-plaski-m7-40-mm-seria-37/",
             "2 – zalecane",
             "Do retuszu końcówek linii, malowania strzałek, numerów i symboli, które nie mogą być nałożone wałkiem. Każde profesjonalne oznakowanie poziome wymaga wykończeń pędzlem."),
            ("Hardy Folia ochronna LDPE\n4×5 m z regranulatu",
             "8,63 zł",
             "https://betondur.eu/produkt/hardy-folia-ochronna-ldpe-4x5-m-z-regranulatu/",
             "2 – zalecane",
             "Ochrona przyległych obszarów (samochodów, urządzeń, ścian) przy malowaniu parkingów i hal. LINER jest farbą szybkoschnącą, ale łatwo ją roznieść na duże powierzchnie. Folia szczególnie ważna gdy klienci to firmy (hale produkcyjne, magazyny)."),
        ]
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # 8. WPC – tarasy kompozytowe
    # ══════════════════════════════════════════════════════════════════════════
    (
        "WPC\nTarasy kompozytowe",
        "Betondur\nImpregnat / Renowator\n/ Środki WPC",
        [
            ("Hardy Wałek malarski Hardstar\n25 cm runo 13 mm Ø48 mm",
             "17,81 zł",
             "https://betondur.eu/produkt/hardy-walek-malarski-hardstar-25-cm-runo-13-mm-o48-mm/",
             "1 – obowiązkowe",
             "Impregnat i renowator WPC nakładany jest wałkiem wzdłuż desek kompozytowych. Runo 13 mm (grubsze) wchodzi w strukturę desek WPC, zapewniając równomierne nasączenie. Wałek 25 cm = szybka praca na całej szerokości tarasu."),
            ("Hardy Gąbka glazurnicza\n23×12×7 cm",
             "24,49 zł",
             "https://betondur.eu/produkt/hardy-gabka-glazurnicza-23x12x7-cm/",
             "1 – obowiązkowe",
             "Środek do plam i środek myjący WPC stosuje się gąbką lub szczotką. Gąbka glazurnicza (twarda, strukturalna) skutecznie usuwa zabrudzenia z rowkowanych desek kompozytowych, docierając do rowków w strukturze WPC. Niezbędna przy czyszczeniu i usuwaniu plam."),
            ("Hardy Kuweta malarska\n26×35 cm (4\")",
             "7,44 zł",
             "https://betondur.eu/produkt/hardy-kuweta-malarska-26x35-cm-4%e2%80%b3/",
             "1 – obowiązkowe",
             "Do nakładania impregnatu i renowatora wałkiem niezbędna jest kuweta. Niska cena = oczywisty cross-sell. Wkładka do kuwety (3,03 zł) jako opcja dodatkowa."),
            ("Hardy Pędzel płaski M7\n40 mm seria 37",
             "5,51 zł",
             "https://betondur.eu/produkt/hardy-pedzel-plaski-m7-40-mm-seria-37/",
             "2 – zalecane",
             "Do wykończenia przy balustradach, ścianach i krawędziach tarasu, gdzie wałek nie dociera. Klienci z tarasami WPC to zazwyczaj właściciele domów dbający o jakość – pędzel pozwala na profesjonalne wykończenie."),
            ("Scley Taśma malarska precyzyjna\nWashi 572, 30 mm × 33 m",
             "9,07 zł",
             "https://betondur.eu/produkt/scley-tasma-malarska-precyzyjna-washi-572-30-mm-x-33-m-do-60-dni/",
             "2 – zalecane",
             "Maskowanie balustrad, słupków i elementów, na które nie powinien trafić impregnat lub renowator. Taśma Washi sprawdza się na zewnątrz (do 60 dni) – co jest istotne przy pracach tarasowych przerywanych przez pogodę."),
        ]
    ),

    # ══════════════════════════════════════════════════════════════════════════
    # 9. ŻYWICA EPOKSYDOWA
    # ══════════════════════════════════════════════════════════════════════════
    (
        "ŻYWICA\nEpoksydowa",
        "Betondur Żywica\nEpoksydowa Bezbarwna\n+ Utwardzacz 800g",
        [
            ("Hardy Szpachelka malarska 8 cm\nstal nierdzewna seria 78",
             "12,51 zł",
             "https://betondur.eu/produkt/hardy-szpachelka-malarska-8-cm-stal-nierdzewna-seria-78/",
             "1 – obowiązkowe",
             "Żywica epoksydowa nakładana i rozprowadzana jest szpachelką lub pacą – to standardowa technika aplikacji. Szpachelka ze stali nierdzewnej (odpornej na chemikalia) jest właściwym narzędziem do żywic. Bez szpachelki nie można równomiernie rozprowadzić żywicy na powierzchni."),
            ("Hardy Pędzel płaski M7\n40 mm seria 37",
             "5,51 zł",
             "https://betondur.eu/produkt/hardy-pedzel-plaski-m7-40-mm-seria-37/",
             "1 – obowiązkowe",
             "Żywicę w narożnikach, krawędziach i na mniejszych detalach nakłada się pędzlem. Pędzel jest też niezbędny do zagruntowania powierzchni przed aplikacją żywicy. Pierwsza warstwa (bonding coat) często nakładana jest właśnie pędzlem."),
            ("Hardy Folia ochronna HDPE\n4×5 m Standard",
             "15,97 zł",
             "https://betondur.eu/produkt/hardy-folia-ochronna-hdpe-4x5-m-standard/",
             "1 – obowiązkowe",
             "Żywica epoksydowa jest ekstremalnie trudna do usunięcia po utwardzeniu (7 dni!). Ochrona podłoża podczas pracy jest krytyczna. HDPE (mocniejsza) jest konieczna – żywica może przebić cieńszą folię LDPE. Najważniejszy cross-sell przy żywicy."),
            ("Scley Taśma malarska precyzyjna\nWashi 572, 48 mm × 33 m",
             "14,78 zł",
             "https://betondur.eu/produkt/scley-tasma-malarska-precyzyjna-washi-572-48-mm-x-33-m-do-60-dni/",
             "1 – obowiązkowe",
             "Żywica epoksydowa stosowana na ograniczonych obszarach (naprawy, elementy dekoracyjne, powłoki) wymaga precyzyjnego maskowania. Szeroka taśma 48 mm zapobiega wyciekaniu żywicy na sąsiednie powierzchnie. Przy żywicy nie ma miejsca na błędy – utwardzonej nie da się usunąć bez szlifowania."),
            ("Hardy Kuweta malarska\n26×35 cm (4\")",
             "7,44 zł",
             "https://betondur.eu/produkt/hardy-kuweta-malarska-26x35-cm-4%e2%80%b3/",
             "2 – zalecane",
             "Do mieszania żywicy z utwardzaczem i do nakładania jej pędzlem/wałkiem. Kuweta jest jednorazowym naczyniem roboczym (żywica utwardza się w pojemniku). Niska cena = klient kupuje kilka wkładek na wypadek dłuższego projektu."),
        ]
    ),
]

# ─── wypełnianie arkusza ───────────────────────────────────────────────────────
row = 3
for cat_name, product_name, accessories in data:
    n_rows = len(accessories)

    # kolumna A – kategoria (scalona)
    ws.merge_cells(start_row=row, start_column=1, end_row=row + n_rows - 1, end_column=1)
    cat_cell = ws.cell(row=row, column=1, value=cat_name)
    cat_cell.fill      = hdr_fill(C_CAT_BG)
    cat_cell.font      = hdr_font(C_HEADER_FONT, size=10)
    cat_cell.alignment = center()
    cat_cell.border    = full_border

    # kolumna B – produkt główny (scalona)
    ws.merge_cells(start_row=row, start_column=2, end_row=row + n_rows - 1, end_column=2)
    prod_cell = ws.cell(row=row, column=2, value=product_name)
    prod_cell.fill      = hdr_fill(C_PROD_BG)
    prod_cell.font      = Font(bold=True, size=9, color="1F3864")
    prod_cell.alignment = center()
    prod_cell.border    = full_border

    # wiersze akcesoriów
    for i, (acc_name, acc_price, acc_url, priority, justification) in enumerate(accessories):
        r = row + i
        bg = C_ACC_BG if i % 2 == 0 else C_ALT_ACC_BG

        # C – akcesorium
        c_cell = ws.cell(row=r, column=3, value=acc_name)
        c_cell.fill = hdr_fill(bg); c_cell.font = Font(size=9, bold=True)
        c_cell.alignment = left(); c_cell.border = full_border

        # D – cena
        d_cell = ws.cell(row=r, column=4, value=acc_price)
        d_cell.fill = hdr_fill(bg); d_cell.font = Font(size=9)
        d_cell.alignment = center(); d_cell.border = full_border

        # E – URL
        e_cell = ws.cell(row=r, column=5, value=acc_url)
        e_cell.fill = hdr_fill(bg); e_cell.font = Font(size=8, color="2E74B5", underline="single")
        e_cell.alignment = left(); e_cell.border = full_border

        # F – priorytet
        if "1" in priority:
            p_fill = PatternFill("solid", fgColor="C6EFCE")
            p_font = Font(size=9, bold=True, color="276221")
        else:
            p_fill = PatternFill("solid", fgColor="FFEB9C")
            p_font = Font(size=9, bold=False, color="9C6500")
        f_cell = ws.cell(row=r, column=6, value=priority)
        f_cell.fill = p_fill; f_cell.font = p_font
        f_cell.alignment = center(); f_cell.border = full_border

        # G – uzasadnienie
        g_cell = ws.cell(row=r, column=7, value=justification)
        g_cell.fill = hdr_fill(C_UZAS_BG); g_cell.font = Font(size=9)
        g_cell.alignment = left(); g_cell.border = full_border

        ws.row_dimensions[r].height = 60

    row += n_rows

# ─────────────────────────────────────────────────────────────────────────────
# ARKUSZ 2 – MATRYCA PRIORYTETÓW (skrócone zestawienie)
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Matryca priorytetów")
ws2.freeze_panes = "B3"

# tytuł
ws2.merge_cells("A1:J1")
t2 = ws2["A1"]
t2.value = "BETONDUR.EU – Matryca priorytetów akcesoriów wg kategorii (widok skrócony)"
t2.fill  = hdr_fill(C_HEADER_BG)
t2.font  = Font(color=C_HEADER_FONT, bold=True, size=12)
t2.alignment = center()
ws2.row_dimensions[1].height = 28

# nagłówki: kategorie jako kolumny
categories_short = [
    "BETON\nEfekt dek.", "RDZA\nEfekt dek.", "OBSYDIAN\nEfekt dek.",
    "HARDFLOOR\nPosadzki", "EKO\nFarba beton", "OUTDOOR\nZewnętrzna",
    "LINER\nZnakowanie", "WPC\nTarasy", "ŻYWICA\nEpoks."
]
ws2.cell(row=2, column=1, value="Akcesorium").fill = hdr_fill(C_HEADER_BG)
ws2.cell(row=2, column=1).font = hdr_font(C_HEADER_FONT)
ws2.cell(row=2, column=1).alignment = center()
ws2.cell(row=2, column=1).border = full_border
ws2.column_dimensions["A"].width = 36

for ci, cat in enumerate(categories_short, start=2):
    cell = ws2.cell(row=2, column=ci, value=cat)
    cell.fill = hdr_fill(C_CAT_BG); cell.font = hdr_font(C_HEADER_FONT, size=9)
    cell.alignment = center(); cell.border = full_border
    ws2.column_dimensions[get_column_letter(ci)].width = 14
ws2.row_dimensions[2].height = 42

# macierz aksesoriów
accessories_matrix = [
    # (nazwa akcesorium, [czy w kategorii: 1=obowiązkowe, 2=zalecane, 0=brak])
    # kolejność kategorii: BETON RDZA OBSYDIAN HARDFLOOR EKO OUTDOOR LINER WPC ŻYWICA
    ("Hardy Paca Stucco wenecka (25/28/35)",   ["1","1","1","0","0","0","0","0","0"]),
    ("Hardy Wałek 25 cm runo 12 mm",            ["1","0","0","1","0","1","0","1","0"]),
    ("Hardy Wałek Hardstar 25 cm runo 13 mm",   ["0","0","1","1","1","0","0","1","0"]),
    ("Hardy Wałek Hardstar 10 cm runo 9 mm",    ["0","0","0","0","0","0","1","0","0"]),
    ("Hardy Zestaw malarski nr 64",             ["0","0","0","1","1","0","0","0","0"]),
    ("Hardy Zestaw malarski nr 68",             ["0","0","0","0","0","1","0","0","0"]),
    ("Hardy Pędzel płaski M7 40 mm",            ["2","2","2","1","1","2","2","2","1"]),
    ("Hardy Gąbka glazurnicza 23×12×7 cm",      ["0","1","0","0","0","0","0","1","0"]),
    ("Hardy Szpachelka 8 cm stal nierdzewna",   ["0","0","0","0","0","0","0","0","1"]),
    ("Hardy Kuweta 26×35 cm (4\")",             ["2","2","0","0","0","0","0","1","2"]),
    ("Hardy Kuweta ręczna 1 l (wałek 15 cm)",   ["0","0","2","0","0","0","0","0","0"]),
    ("Hardy Kratka malarska do wiadra",         ["0","0","0","2","0","0","0","0","0"]),
    ("Hardy Folia HDPE 4×5 m Standard",         ["1","0","1","0","0","0","0","0","1"]),
    ("Hardy Folia LDPE 4×5 m z regranulatu",    ["0","1","0","2","2","0","2","0","0"]),
    ("Scley Taśma Washi 572 – 30 mm × 33 m",   ["0","1","0","0","2","0","0","2","0"]),
    ("Scley Taśma Washi 572 – 48 mm × 33 m",   ["1","0","1","1","0","1","1","0","1"]),
    ("Scley Tektura malarska falista 1×15 m",   ["0","0","0","0","0","2","0","0","0"]),
]

fill_1 = PatternFill("solid", fgColor="C6EFCE")  # zielony – obowiązkowe
fill_2 = PatternFill("solid", fgColor="FFEB9C")  # żółty – zalecane
fill_0 = PatternFill("solid", fgColor="F5F5F5")  # szary – brak
font_1 = Font(bold=True, size=9, color="276221")
font_2 = Font(bold=False, size=9, color="9C6500")
font_0 = Font(size=9, color="BBBBBB")

labels = {"1": "● obowiązkowe", "2": "○ zalecane", "0": "—"}

for ri, (acc_name, presence) in enumerate(accessories_matrix, start=3):
    ws2.cell(row=ri, column=1, value=acc_name).font = Font(size=9, bold=True)
    ws2.cell(row=ri, column=1).alignment = left()
    ws2.cell(row=ri, column=1).border = full_border
    ws2.cell(row=ri, column=1).fill = hdr_fill("EBF3FB" if ri % 2 == 0 else "FFFFFF")

    for ci, val in enumerate(presence, start=2):
        cell = ws2.cell(row=ri, column=ci, value=labels[val])
        if val == "1":
            cell.fill = fill_1; cell.font = font_1
        elif val == "2":
            cell.fill = fill_2; cell.font = font_2
        else:
            cell.fill = fill_0; cell.font = font_0
        cell.alignment = center()
        cell.border = full_border
    ws2.row_dimensions[ri].height = 22

# ─────────────────────────────────────────────────────────────────────────────
# ARKUSZ 3 – LOGIKA WDROŻENIA (rekomendacje techniczne)
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Logika wdrożenia")

ws3.merge_cells("A1:D1")
t3 = ws3["A1"]
t3.value = "BETONDUR.EU – Logika wdrożenia zestawów w WooCommerce"
t3.fill = hdr_fill(C_HEADER_BG); t3.font = Font(color=C_HEADER_FONT, bold=True, size=12)
t3.alignment = center(); ws3.row_dimensions[1].height = 28

headers3 = ["Aspekt", "Rekomendacja", "Uzasadnienie biznesowe", "Priorytet wdrożenia"]
widths3  = [28, 42, 56, 20]
for ci, (h, w) in enumerate(zip(headers3, widths3), start=1):
    cell = ws3.cell(row=2, column=ci, value=h)
    cell.fill = hdr_fill(C_CAT_BG); cell.font = hdr_font(C_HEADER_FONT)
    cell.alignment = center(); cell.border = full_border
    ws3.column_dimensions[get_column_letter(ci)].width = w
ws3.row_dimensions[2].height = 28

impl_data = [
    ("Tryb wyświetlania\nakcesoriów",
     "Popup / drawer po dodaniu do koszyka\nlub sekcja 'Uzupelnij swoj zestaw'\nna stronie produktu (before checkout)",
     "Popup po add-to-cart ma najwyższy CTR (do 35%). Sekcja na stronie produktu buduje zaufanie i edukuje klienta PRZED decyzją zakupową – szczególnie ważne dla produktów dekoracyjnych, gdzie klient musi wiedzieć, że potrzebuje pacy weneckiej.",
     "1 – WYSOKI"),
    ("Liczba sugerowanych\nakcesoriów",
     "Max 3-4 akcesoria w popup\n(tylko priorytet 1)\nNa stronie produktu: do 6",
     "Paradoks wyboru – zbyt wiele opcji blokuje decyzję. W popup pokaż tylko absolutnie niezbędne akcesoria (priorytet 1). Na stronie produktu można pokazać pełen zestaw z podziałem na 'Niezbędne' i 'Polecane'.",
     "1 – WYSOKI"),
    ("Pakietowanie\n(Bundle discount)",
     "Oferuj rabat 5-10% przy dodaniu\nmin. 2 akcesoriów do produktu głównego\nLub: 'Kup zestaw i zaoszczędź X zł'",
     "Bundle pricing zwiększa AOV (Average Order Value) o 20-30%. Przy produktach Betondur (mass dekoracyjne 200-400 zł) dodanie akcesoriów za 50-100 zł z rabatkiem jest bardzo atrakcyjne i realistyczne dla klienta.",
     "1 – WYSOKI"),
    ("Kolejność\nakcesoriów w zestawie",
     "1. Paca wenecka (dla efektów dek.)\n2. Wałek malarski\n3. Taśma Washi\n4. Folia ochronna\n5. Pędzel\n6. Kuweta",
     "Sortowanie wg ważności procesu nakładania (od narzędzia do aplikacji, przez maskowanie, po ochronę). Klient myśli sekwencyjnie – 'co mi potrzeba najpierw'. Psychologicznie: zacznij od najdroższego akcesorium (paca ~30 zł), a tańsze (taśma, pędzel) dochodzą naturalnie.",
     "2 – ŚREDNI"),
    ("Tekst CTA\ndla zestawów",
     "Efekty dekoracyjne: 'Dobierz narzędzia\ndo efektu BETON/RDZA/OBSYDIAN'\nFarby techniczne: 'Co będziesz potrzebować?'\nWPC: 'Kompletny zestaw do tarasu'",
     "Personalizacja CTA wg linii produktowej podnosi CTR o 15-25% vs. generyczne 'Kup też'. Klient kupujący efekt dekoracyjny jest w innym mindset niż kupujący farbę do garażu – komunikat musi to odzwierciedlać.",
     "2 – ŚREDNI"),
    ("Cross-sell przy\ngotowych zestawach\n(BETON/RDZA/OBSYDIAN\nCompletny Zestaw)",
     "Do gotowych zestawów proponuj:\n1. Paca wenecka (priorytet 1)\n2. Taśma Washi 48mm (priorytet 1)\n3. Folia HDPE (priorytet 1)\nBEZ wałka i lakieru (już w zestawie)",
     "Klient kupujący kompletny zestaw już ma produkt chemiczny w komplecie. Jego potrzeba to NARZĘDZIA i OCHRONA. Nie pokazuj produktów, które już kupił – to psuje UX i obniża zaufanie. Filtruj akcesoria wg zawartości zestawu.",
     "1 – WYSOKI"),
    ("Sezonowość\ni priorytety",
     "Wiosna/Lato: promuj zestawy OUTDOOR\ni WPC (tarasy)\nJesień/Zima: efekty dekoracyjne\n(remonty wewnętrzne)",
     "E-com w branży budowlanej ma silną sezonowość. Prace zewnętrzne (tarasy, balkony, parking) – marzec-wrzesień. Remonty wnętrz i efekty dekoracyjne – październik-marzec. Dostosuj kolejność sugestii w pluginie wg sezonu.",
     "3 – NISKI"),
    ("Dane do optymalizacji",
     "Monitoruj:\n- CTR zestawów wg kategorii\n- Konwersję add-to-cart → zakup\n- AOV z/bez aksesoriów\n- Najczęściej dodawane akcesoria",
     "Po 4-8 tygodniach masz dane do optymalizacji. Zestawy wg danych zastąpią zestawy wg intuicji. Testuj A/B: popup vs. sekcja na stronie produktu, 3 vs. 5 akcesoriów, z rabatem vs. bez.",
     "2 – ŚREDNI"),
]

for ri, row_data in enumerate(impl_data, start=3):
    bg = "EBF3FB" if ri % 2 == 0 else "FFFFFF"
    for ci, val in enumerate(row_data, start=1):
        cell = ws3.cell(row=ri, column=ci, value=val)
        cell.fill = hdr_fill(bg)
        if ci == 4:
            if "1" in val:
                cell.fill = fill_1; cell.font = font_1
            elif "2" in val:
                cell.fill = fill_2; cell.font = font_2
            else:
                cell.fill = PatternFill("solid", fgColor="FCE4D6"); cell.font = Font(size=9, color="C55A11")
        else:
            cell.font = Font(size=9)
        cell.alignment = left()
        cell.border = full_border
    ws3.row_dimensions[ri].height = 55

# zapis
output_path = "/Users/agencjaaqq/ai-lab/Betondur_Zestawy_Akcesoriow.xlsx"
wb.save(output_path)
print(f"Zapisano: {output_path}")
