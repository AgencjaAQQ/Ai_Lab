import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── kolory ──────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E74B5"
LIGHT_BLUE  = "D6E4F0"
PALE_BLUE   = "EBF3FB"
ALT_WHITE   = "F8FBFF"
GREEN_BG    = "C6EFCE"
GREEN_FG    = "276221"
YELLOW_BG   = "FFEB9C"
YELLOW_FG   = "9C6500"
CREAM       = "FFFDE7"
BORDER_COL  = "B0C4D8"

# produkty – kolory pasów (naprzemiennie na grupę produktową)
GROUP_COLORS = [
    ("E8F4FD", "D0E8F7"),  # jasny niebieski
    ("EDF7ED", "D5EED5"),  # jasny zielony
    ("FFF8E1", "FDEEC0"),  # jasny żółty
    ("F3E5F5", "E1BEE7"),  # jasny fiolet
    ("FCE4EC", "F8BBD0"),  # jasny różowy
    ("E0F2F1", "B2DFDB"),  # jasny morski
    ("FBE9E7", "FFCCBC"),  # jasny pomarańcz
    ("F5F5F5", "E0E0E0"),  # szary
]

def fill(c): return PatternFill("solid", fgColor=c)
def font(c="000000", bold=False, sz=9): return Font(color=c, bold=bold, size=sz)
def left(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(style="thin", color=BORDER_COL)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ══════════════════════════════════════════════════════════════════════════════
# DANE — (nr_kat, kategoria_sklepu, produkt, akcesorium, cena_akc, priorytet, uzasadnienie)
# priorytet: "1" = obowiązkowe, "2" = zalecane
# ══════════════════════════════════════════════════════════════════════════════
rows = [

    # ── 1. BETON efekt dekoracyjny ──────────────────────────────────────────
    (1,"Efekty dekoracyjne","Betondur BETON efekt dekoracyjny do ścian",
     "Hardy Paca Stucco (wenecka) seria 35, 24x9 cm, uchwyt 2K","34,51 zl",
     "1 - obowiazkowe",
     "Masa BETON nakladana jest wylacznie paca wenecka warstwa ~1 mm. Bez odpowiedniej pacy niemoliwe jest uzyskanie efektu betonu. Seria 35 (24x9 cm) – optymalny rozmiar do scian."),
    (1,"Efekty dekoracyjne","Betondur BETON efekt dekoracyjny do ścian",
     "Hardy Walek malarski 25 cm, runo 12 mm","20,06 zl",
     "1 - obowiazkowe",
     "Podklad i lakier zabezpieczajacy naklada sie walkiem. Runo 12 mm = rowne krycie na scianach bez nadmiernego chlonienia materialu."),
    (1,"Efekty dekoracyjne","Betondur BETON efekt dekoracyjny do ścian",
     "Scley Tasma malarska Washi 572, 48 mm x 33 m","14,78 zl",
     "1 - obowiazkowe",
     "Precyzyjne maskowanie narozninkow, listew, okien. Szerokose 48 mm = lepsza ochrona przy nakladaniu masy paca. Washi nie niszczy podloza (do 60 dni)."),
    (1,"Efekty dekoracyjne","Betondur BETON efekt dekoracyjny do ścian",
     "Hardy Folia ochronna HDPE 4x5 m Standard","15,97 zl",
     "1 - obowiazkowe",
     "Masa BETON jest trudna do usuniecia z podlog i mebli. Folia HDPE (mocniejsza) chroni cale pomieszczenie. Klienci bez doswiadczenia – folia zapobiega kosztownym uszkodzeniom."),
    (1,"Efekty dekoracyjne","Betondur BETON efekt dekoracyjny do ścian",
     "Hardy Kuweta malarska 26x35 cm (4\")","7,44 zl",
     "2 - zalecane",
     "Do nakladania lakieru i podkladu walkiem. Niska cena = oczywisty cross-sell. Warto oferowac z wkladka do kuwety (3,03 zl)."),
    (1,"Efekty dekoracyjne","Betondur BETON efekt dekoracyjny do ścian",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "2 - zalecane",
     "Narozniki scian i miejsca przy listwach wymagaja pedzelka. Niska cena podnosi AOV bez oporu klienta."),

    # ── 2. OBSYDIAN efekt dekoracyjny ───────────────────────────────────────
    (2,"Efekty dekoracyjne","Betondur OBSYDIAN efekt dekoracyjny do scian (czarny)",
     "Hardy Paca Stucco (wenecka) seria 35, 24x9 cm, uchwyt 2K","34,51 zl",
     "1 - obowiazkowe",
     "OBSYDIAN to masa strukturalna z ziarnistoscia ~1 mm, nakladana paca wenecka. Wieksza paca = efektywniejsze tworzenie wglebien (efekt kamienia wulkanicznego). Bez pacy produkt jest niemozliwy do naloenia."),
    (2,"Efekty dekoracyjne","Betondur OBSYDIAN efekt dekoracyjny do scian (czarny)",
     "Hardy Walek malarski Hardstar 25 cm, runo 13 mm","17,81 zl",
     "1 - obowiazkowe",
     "Podklad OBSYDIAN naklada sie walkiem. Runo 13 mm (grubsze) lepiej sprawdza sie przy podkladach akrylowych na chropowatych powierzchniach."),
    (2,"Efekty dekoracyjne","Betondur OBSYDIAN efekt dekoracyjny do scian (czarny)",
     "Scley Tasma malarska Washi 572, 48 mm x 33 m","14,78 zl",
     "1 - obowiazkowe",
     "OBSYDIAN to efekt premium – klienci dbaja o detal. Szersza tasma 48 mm chroni wiecej powierzchni przy nakladaniu grubej, strukturalnej masy."),
    (2,"Efekty dekoracyjne","Betondur OBSYDIAN efekt dekoracyjny do scian (czarny)",
     "Hardy Folia ochronna HDPE 4x5 m Standard","15,97 zl",
     "1 - obowiazkowe",
     "Czarna masa OBSYDIAN bardzo trudna do usuniecia. HDPE (mocna folia) chroni podloge i meble. Klienci premium docenia profesjonalne podejscie."),
    (2,"Efekty dekoracyjne","Betondur OBSYDIAN efekt dekoracyjny do scian (czarny)",
     "Hardy Kuweta reczna malarska 1 l (do walka 15 cm)","27,83 zl",
     "2 - zalecane",
     "Kuweta reczna (z raczka) jest wygodniejsza do pracy z lakierem/podkladem OBSYDIAN przy trudno dostepnych miejscach. Model premium dopasowany do klienta kupujacego produkt z wyzszej polki."),
    (2,"Efekty dekoracyjne","Betondur OBSYDIAN efekt dekoracyjny do scian (czarny)",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "2 - zalecane",
     "Narozniki i krawedzie przy czarnym efekcie – kazda nierowniomierna krawedz bedzie widoczna. Pedzel do dokladnych wykonczen."),

    # ── 3. RDZA efekt dekoracyjny ────────────────────────────────────────────
    (3,"Efekty dekoracyjne","Betondur RDZA efekt dekoracyjny do scian",
     "Hardy Paca Stucco (wenecka) seria 25, 20x8 cm, uchwyt drewniany","29,13 zl",
     "1 - obowiazkowe",
     "RDZA nakladana paca wenecka nieregularnymi ruchami – to fundamentalne narzedzie. Mniejsza seria 25 (20x8 cm) daje bardziej chaotyczny, naturalny efekt rdzy."),
    (3,"Efekty dekoracyjne","Betondur RDZA efekt dekoracyjny do scian",
     "Hardy Gabka glazurnicza 23x12x7 cm","24,49 zl",
     "1 - obowiazkowe",
     "Aktywator do efektu rdzy nalezy rozprowadzic gabka w nieregularny sposob – to tym wlasnie uzyskuje sie autentyczny efekt korozji. Gabka strukturalna dociera do tekstury masy."),
    (3,"Efekty dekoracyjne","Betondur RDZA efekt dekoracyjny do scian",
     "Scley Tasma malarska Washi 572, 30 mm x 33 m","9,07 zl",
     "1 - obowiazkowe",
     "RDZA czesto nakladana jako accent wall (wybrany fragment). Precyzyjne granice efektu sa kluczowe. Aktywator jest agresywny chemicznie – tasma chroni sasiendne powierzchnie."),
    (3,"Efekty dekoracyjne","Betondur RDZA efekt dekoracyjny do scian",
     "Hardy Folia ochronna LDPE 4x5 m z regranulatu","8,63 zl",
     "1 - obowiazkowe",
     "Aktywator do rdzy moze trwale zabawic podloge lub meble. Folia LDPE (tansza) wystarcza do ochrony podlogi podczas nakladania."),
    (3,"Efekty dekoracyjne","Betondur RDZA efekt dekoracyjny do scian",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "2 - zalecane",
     "Pedzel do podkladu i lakieru przy kraweddziach. Mozna tez uzyc do rozprofilowania efektu rdzy przy listwach."),
    (3,"Efekty dekoracyjne","Betondur RDZA efekt dekoracyjny do scian",
     "Hardy Kuweta malarska 26x35 cm (4\")","7,44 zl",
     "2 - zalecane",
     "Do nakladania podkladu i lakieru walkiem. Uzupelnia zestaw narzedzi."),

    # ── 4. BETON kompletny zestaw ────────────────────────────────────────────
    (4,"Zestawy","Betondur BETON kompletny zestaw (podklad + masa + lakier)",
     "Hardy Paca Stucco (wenecka) seria 35, 24x9 cm, uchwyt 2K","34,51 zl",
     "1 - obowiazkowe",
     "Zestaw zawiera chemia – klient nadal potrzebuje narzedzi. Paca wenecka to absolutnie niezbedne narzedzie do nakladania masy BETON. Kupujacy gotowy zestaw to czesciej poczatkujacy – potrzebuje kompletnych wskazowek."),
    (4,"Zestawy","Betondur BETON kompletny zestaw (podklad + masa + lakier)",
     "Hardy Walek malarski 25 cm, runo 12 mm","20,06 zl",
     "1 - obowiazkowe",
     "Podklad i lakier z zestawu naklada sie walkiem. Bez walka klient nie moze zastosowac 2 z 3 skladnikow zestawu."),
    (4,"Zestawy","Betondur BETON kompletny zestaw (podklad + masa + lakier)",
     "Scley Tasma malarska Washi 572, 48 mm x 33 m","14,78 zl",
     "1 - obowiazkowe",
     "Maskowanie przy pracach dekoracyjnych – obowiazkowe niezaleznie czy kupujemy chem osobno czy w zestawie."),
    (4,"Zestawy","Betondur BETON kompletny zestaw (podklad + masa + lakier)",
     "Hardy Folia ochronna HDPE 4x5 m Standard","15,97 zl",
     "1 - obowiazkowe",
     "Ochrona podlogi i mebli. Klient kupujacy kompletny zestaw (wyzszy koszt) to osoba ktora zainwestowala w produkt – tym bardziej warto chronic powierzchnie przed zniszczeniem."),
    (4,"Zestawy","Betondur BETON kompletny zestaw (podklad + masa + lakier)",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "2 - zalecane",
     "Wykончenia naroznikow pedzlem. Niska cena wzglem wartosci zestawu."),
    (4,"Zestawy","Betondur BETON kompletny zestaw (podklad + masa + lakier)",
     "Hardy Kuweta malarska 26x35 cm (4\")","7,44 zl",
     "2 - zalecane",
     "Do nakladania podkladu i lakieru walkiem. Podstawowe akcesorium malarskie."),

    # ── 5. OBSYDIAN kompletny zestaw ─────────────────────────────────────────
    (5,"Zestawy","Betondur OBSYDIAN kompletny zestaw",
     "Hardy Paca Stucco (wenecka) seria 35, 24x9 cm, uchwyt 2K","34,51 zl",
     "1 - obowiazkowe",
     "Masa OBSYDIAN nakladana paca wenecka – narzedzie niezbedne, chociaz zestaw zawiera caly zestaw chemiczny. Seria 35 = optymalny rozmiar dla strukturalnej masy."),
    (5,"Zestawy","Betondur OBSYDIAN kompletny zestaw",
     "Hardy Walek malarski Hardstar 25 cm, runo 13 mm","17,81 zl",
     "1 - obowiazkowe",
     "Do nakladania podkladu i lakieru z zestawu. Runo 13 mm dla produktow akrylowych."),
    (5,"Zestawy","Betondur OBSYDIAN kompletny zestaw",
     "Scley Tasma malarska Washi 572, 48 mm x 33 m","14,78 zl",
     "1 - obowiazkowe",
     "Maskowanie przy pracach efektowych. Szersza tasma do grubszej masy strukturalnej."),
    (5,"Zestawy","Betondur OBSYDIAN kompletny zestaw",
     "Hardy Folia ochronna HDPE 4x5 m Standard","15,97 zl",
     "1 - obowiazkowe",
     "Czarna masa OBSYDIAN jest szczegolnie widoczna na podlogach i meblach. HDPE chroni skutecznie."),
    (5,"Zestawy","Betondur OBSYDIAN kompletny zestaw",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "2 - zalecane",
     "Narozniki i wykончenia detali. Efekt premium wymaga precyzji."),

    # ── 6. RDZA kompletny zestaw ─────────────────────────────────────────────
    (6,"Zestawy","Betondur RDZA kompletny zestaw",
     "Hardy Paca Stucco (wenecka) seria 25, 20x8 cm, uchwyt drewniany","29,13 zl",
     "1 - obowiazkowe",
     "Masa RDZA z zestawu nakladana paca wenecka. Mniejsza seria 25 daje bardziej nieregularny, naturalny efekt rdzy."),
    (6,"Zestawy","Betondur RDZA kompletny zestaw",
     "Hardy Gabka glazurnicza 23x12x7 cm","24,49 zl",
     "1 - obowiazkowe",
     "Aktywator (zawarty w zestawie) naklada sie gabka. Bez gabki klient nie moze prawidlowo zastosowac aktywatora – to kluczowy etap tworzenia efektu rdzy."),
    (6,"Zestawy","Betondur RDZA kompletny zestaw",
     "Scley Tasma malarska Washi 572, 30 mm x 33 m","9,07 zl",
     "1 - obowiazkowe",
     "Maskowanie krawedzi i sasiednich powierzchni – aktywator agresywny chemicznie, nalezy chronic otoczenie."),
    (6,"Zestawy","Betondur RDZA kompletny zestaw",
     "Hardy Folia ochronna LDPE 4x5 m z regranulatu","8,63 zl",
     "1 - obowiazkowe",
     "Aktywator moze trwale zabarwic podloge. Folia LDPE wystarczy do ochrony podczas nakladania."),
    (6,"Zestawy","Betondur RDZA kompletny zestaw",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "2 - zalecane",
     "Narozniki i krawedzie – pedzel do wykonczenia."),

    # ── 7. BETON lakier bezbarwny ─────────────────────────────────────────────
    (7,"Lakiery","Betondur BETON lakier zabezpieczajacy (matowy, bezbarwny)",
     "Hardy Walek malarski 25 cm, runo 12 mm","20,06 zl",
     "1 - obowiazkowe",
     "Lakier do scian naklada sie walkiem. Runo 12 mm = rowne, cienkie krycie bez zaciekow. Klient kupujacy sam lakier musi miec wlasne narzedzia – warto zaproponowac."),
    (7,"Lakiery","Betondur BETON lakier zabezpieczajacy (matowy, bezbarwny)",
     "Hardy Kuweta malarska 26x35 cm (4\")","7,44 zl",
     "1 - obowiazkowe",
     "Do walka potrzebna kuweta. Niska cena, oczywisty cross-sell."),
    (7,"Lakiery","Betondur BETON lakier zabezpieczajacy (matowy, bezbarwny)",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "1 - obowiazkowe",
     "Narozniki i miejsca przy listwach. Lakier musi byc rownomierny – pedzel do precyzji krawedzi."),
    (7,"Lakiery","Betondur BETON lakier zabezpieczajacy (matowy, bezbarwny)",
     "Scley Tasma malarska Washi 572, 30 mm x 33 m","9,07 zl",
     "2 - zalecane",
     "Maskowanie przy aplikacji lakieru – chroni elementy, ktore nie moga byc polakierowane (listwy metalowe, okna)."),

    # ── 8. BETON lakier czarny ────────────────────────────────────────────────
    (8,"Lakiery","Betondur BETON lakier zabezpieczajacy (matowy, czarny)",
     "Hardy Walek malarski 25 cm, runo 12 mm","20,06 zl",
     "1 - obowiazkowe",
     "Lakier czarny naklada sie walkiem – rowne krycie jest kluczowe, poniewaz czarny kolor uwydatnia kazda nierownosc aplikacji."),
    (8,"Lakiery","Betondur BETON lakier zabezpieczajacy (matowy, czarny)",
     "Hardy Kuweta malarska 26x35 cm (4\")","7,44 zl",
     "1 - obowiazkowe",
     "Niezbedna do pracy z walkiem. Wkladka do kuwety (3,03 zl) – warto dodac jako opcje."),
    (8,"Lakiery","Betondur BETON lakier zabezpieczajacy (matowy, czarny)",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "1 - obowiazkowe",
     "Czarny lakier przy kraweddziach – kazdy blad widoczny. Pedzel do precyzyjnych wykonczen."),
    (8,"Lakiery","Betondur BETON lakier zabezpieczajacy (matowy, czarny)",
     "Scley Tasma malarska Washi 572, 48 mm x 33 m","14,78 zl",
     "2 - zalecane",
     "Szersza tasma (48 mm) przy czarnym lakierze – wieksza ochrona sasiednich powierzchni, ktore by sie widocznie zabrudzily."),

    # ── 9. Lakier ogolny ──────────────────────────────────────────────────────
    (9,"Lakiery","Betondur lakier zabezpieczajacy do scian (matowy, bezbarwny) – ogolny",
     "Hardy Walek malarski 25 cm, runo 12 mm","20,06 zl",
     "1 - obowiazkowe",
     "Lakier ogolny do scian – nakladany walkiem na duze powierzchnie. Runo 12 mm rowne krycie."),
    (9,"Lakiery","Betondur lakier zabezpieczajacy do scian (matowy, bezbarwny) – ogolny",
     "Hardy Kuweta malarska 26x35 cm (4\")","7,44 zl",
     "1 - obowiazkowe",
     "Niezbedna do nakladania walkiem. Cross-sell oczywisty."),
    (9,"Lakiery","Betondur lakier zabezpieczajacy do scian (matowy, bezbarwny) – ogolny",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "1 - obowiazkowe",
     "Wykończenia krawedzi i naroznikow."),
    (9,"Lakiery","Betondur lakier zabezpieczajacy do scian (matowy, bezbarwny) – ogolny",
     "Scley Tasma malarska Washi 572, 30 mm x 33 m","9,07 zl",
     "2 - zalecane",
     "Maskowanie elementow przy malowaniu scian lakierem."),

    # ── 10. RDZA lakier ───────────────────────────────────────────────────────
    (10,"Lakiery","Betondur RDZA lakier zabezpieczajacy do scian (matowy, bezbarwny)",
     "Hardy Walek malarski 25 cm, runo 12 mm","20,06 zl",
     "1 - obowiazkowe",
     "Lakier RDZA naklada sie walkiem po 24h od aktywatora. Klient ktory kupuje sam lakier uzupelniajacy musi miec wlasny walek lub go kupic."),
    (10,"Lakiery","Betondur RDZA lakier zabezpieczajacy do scian (matowy, bezbarwny)",
     "Hardy Kuweta malarska 26x35 cm (4\")","7,44 zl",
     "1 - obowiazkowe",
     "Niezbedna do walka."),
    (10,"Lakiery","Betondur RDZA lakier zabezpieczajacy do scian (matowy, bezbarwny)",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "2 - zalecane",
     "Wykončenia krawedzi przy efekcie rdzy."),

    # ── 11. BETON podklad ─────────────────────────────────────────────────────
    (11,"Podklady","Betondur BETON podklad pod efekt dekoracyjny (szary)",
     "Hardy Walek malarski 25 cm, runo 12 mm","20,06 zl",
     "1 - obowiazkowe",
     "Podklad naklada sie walkiem na sciane przed masa BETON. Klient kupujacy sam podklad potrzebuje walka do jego aplikacji."),
    (11,"Podklady","Betondur BETON podklad pod efekt dekoracyjny (szary)",
     "Hardy Kuweta malarska 26x35 cm (4\")","7,44 zl",
     "1 - obowiazkowe",
     "Do nakladania walkiem – niezbedna."),
    (11,"Podklady","Betondur BETON podklad pod efekt dekoracyjny (szary)",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "1 - obowiazkowe",
     "Narozniki i krawedzie – gruntowanie musi byc rowniemierne wszdzie, inaczej masa dekoracyjna bedzie sie roznic kolorystycznie."),
    (11,"Podklady","Betondur BETON podklad pod efekt dekoracyjny (szary)",
     "Scley Tasma malarska Washi 572, 30 mm x 33 m","9,07 zl",
     "2 - zalecane",
     "Maskowanie przy gruntowaniu – chroni listwy i powierzchnie nieprzeznaczone do podkladu."),
    (11,"Podklady","Betondur BETON podklad pod efekt dekoracyjny (szary)",
     "Hardy Folia ochronna LDPE 4x5 m z regranulatu","8,63 zl",
     "2 - zalecane",
     "Ochrona podlogi podczas gruntowania. LDPE wystarczy – podklad latwiej zmyc niz mase dekoracyjna."),

    # ── 12. OBSYDIAN podklad ──────────────────────────────────────────────────
    (12,"Podklady","Betondur OBSYDIAN podklad pod masy dekoracyjne",
     "Hardy Walek malarski Hardstar 25 cm, runo 13 mm","17,81 zl",
     "1 - obowiazkowe",
     "Podklad OBSYDIAN naklada sie walkiem. Runo 13 mm lepiej wnika w chropowate podloza betonowe."),
    (12,"Podklady","Betondur OBSYDIAN podklad pod masy dekoracyjne",
     "Hardy Kuweta malarska 26x35 cm (4\")","7,44 zl",
     "1 - obowiazkowe",
     "Niezbedna do walka."),
    (12,"Podklady","Betondur OBSYDIAN podklad pod masy dekoracyjne",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "1 - obowiazkowe",
     "Narozniki i trudno dostepne miejsca."),
    (12,"Podklady","Betondur OBSYDIAN podklad pod masy dekoracyjne",
     "Scley Tasma malarska Washi 572, 30 mm x 33 m","9,07 zl",
     "2 - zalecane",
     "Maskowanie przy gruntowaniu."),

    # ── 13. RDZA podklad ──────────────────────────────────────────────────────
    (13,"Podklady","Betondur RDZA podklad pod efekt dekoracyjny (szary)",
     "Hardy Walek malarski 25 cm, runo 12 mm","20,06 zl",
     "1 - obowiazkowe",
     "Podklad RDZA naklada sie walkiem przed masa. Klient kupujacy sam podklad potrzebuje walka."),
    (13,"Podklady","Betondur RDZA podklad pod efekt dekoracyjny (szary)",
     "Hardy Kuweta malarska 26x35 cm (4\")","7,44 zl",
     "1 - obowiazkowe",
     "Niezbedna do walka."),
    (13,"Podklady","Betondur RDZA podklad pod efekt dekoracyjny (szary)",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "1 - obowiazkowe",
     "Krawedzie i narozniki – rowne gruntowanie to podstawa dobrego efektu RDZA."),

    # ── 14. EKO farba do betonu ───────────────────────────────────────────────
    (14,"Farby","Betondur EKO farba do betonu (sciany wewn./zewn.)",
     "Hardy Zestaw malarski nr 64 (walek Hardstar 25 cm, runo 13 mm)","47,45 zl",
     "1 - obowiazkowe",
     "EKO to farba do piwnic, gazazy, koltlowni – duze powierzchnie scian. Zestaw nr 64 (kompletny: trzonek + walek + kuweta) to wszystko w jednym. Runo 13 mm dobrze wchodzi w chropowate powierzchnie betonowe."),
    (14,"Farby","Betondur EKO farba do betonu (sciany wewn./zewn.)",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "1 - obowiazkowe",
     "Narozniki, laczenia ze stropem, slupki – walkiem nie da sie wymalowac tych miejsc. Pedzel niezbedny."),
    (14,"Farby","Betondur EKO farba do betonu (sciany wewn./zewn.)",
     "Scley Tasma malarska Washi 572, 30 mm x 33 m","9,07 zl",
     "2 - zalecane",
     "Maskowanie okien, rur, instalacji w piwnicach i gazazach. Tansza wersja (30 mm) pasuje do budzetowejodfarby technicznej."),
    (14,"Farby","Betondur EKO farba do betonu (sciany wewn./zewn.)",
     "Hardy Folia ochronna LDPE 4x5 m z regranulatu","8,63 zl",
     "2 - zalecane",
     "Ochrona podlog i urzadzen podczas malowania sciann. LDPE wystarczy."),

    # ── 15. HARDFLOOR ─────────────────────────────────────────────────────────
    (15,"Farby","Betondur HARDFLOOR farba do posadzek i betonu",
     "Hardy Zestaw malarski nr 64 (walek Hardstar 25 cm, runo 13 mm)","47,45 zl",
     "1 - obowiazkowe",
     "HARDFLOOR na duze powierzchnie posadzek – garaz, magazyn, piwnica. Zestaw nr 64 to kompletny zestaw do malowania podlogi. Runo 13 mm dla gestyc farb posadzkowych."),
    (15,"Farby","Betondur HARDFLOOR farba do posadzek i betonu",
     "Scley Tasma malarska Washi 572, 48 mm x 33 m","14,78 zl",
     "1 - obowiazkowe",
     "Przy malowaniu posadzek konieczne maskowanie krawedzi przy scianach, progach, slupkach. Szersza 48 mm = lepsza ochrona przy wazku."),
    (15,"Farby","Betondur HARDFLOOR farba do posadzek i betonu",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "1 - obowiazkowe",
     "Podbitka (naroznik podlogi przy scianie) malowana pedzelkiem, nie walkiem. Klient nie wymaluje podlogi bez pedzelka."),
    (15,"Farby","Betondur HARDFLOOR farba do posadzek i betonu",
     "Hardy Folia ochronna LDPE 4x5 m z regranulatu","8,63 zl",
     "2 - zalecane",
     "Ochrona scian i mebli przy malowaniu posadzki w piwnicy lub garazu."),
    (15,"Farby","Betondur HARDFLOOR farba do posadzek i betonu",
     "Hardy Kratka malarska do wiadra 27x32 cm","4,53 zl",
     "2 - zalecane",
     "Kratka do wiadra praktyczniejsza od kuwety przy duzych posadzkach – wiadro mozna przesuwac po podloze. Niska cena = latwa decyzja."),

    # ── 16. LINER ─────────────────────────────────────────────────────────────
    (16,"Farby","Betondur LINER (farba do znakowania poziomego)",
     "Scley Tasma malarska Washi 572, 48 mm x 33 m","14,78 zl",
     "1 - obowiazkowe",
     "Linie parkingowe i pasy wymagaja idealnie prostych krawedzi. Tasma 48 mm to standard przy znakowaniu poziomym – wyznacza szerokosc pasa i chroni otoczenie."),
    (16,"Farby","Betondur LINER (farba do znakowania poziomego)",
     "Hardy Walek malarski Hardstar 10 cm, runo 9 mm","11,87 zl",
     "1 - obowiazkowe",
     "Maly walek 10 cm idealny do malowania wskich linii parkingowych – miesci sie miedzy dwiema tasmami. Runo 9 mm (cienkie) = rowne naniesienie gestej farby LINER na gladkich powierzchniach."),
    (16,"Farby","Betondur LINER (farba do znakowania poziomego)",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "2 - zalecane",
     "Do retuszu koncowek linii, malowania strzalek, numerow i symboli."),
    (16,"Farby","Betondur LINER (farba do znakowania poziomego)",
     "Hardy Folia ochronna LDPE 4x5 m z regranulatu","8,63 zl",
     "2 - zalecane",
     "Ochrona sasiednich obszarow przy malowaniu parkingów i hal produkcyjnych."),

    # ── 17. OUTDOOR ───────────────────────────────────────────────────────────
    (17,"Farby","Betondur OUTDOOR powloka akrylowa na beton (zewnatrz)",
     "Hardy Zestaw malarski nr 68 (walek Hardmax 18 cm, runo 18 mm)","30,20 zl",
     "1 - obowiazkowe",
     "Runo 18 mm (grube) jest szczegolnie skuteczne na chropowatych powierzchniach zewnetrznych (beton, asfalt). Zestaw kompletny z trzonkiem i kuweta."),
    (17,"Farby","Betondur OUTDOOR powloka akrylowa na beton (zewnatrz)",
     "Hardy Walek malarski 25 cm, runo 12 mm","20,06 zl",
     "1 - obowiazkowe",
     "Wiekszy walek 25 cm do gladszych i wiekszych powierzchni zewnetrznych. Komplementarny z nr 68 – razem pokrywaja wszystkie typy nawierzchni."),
    (17,"Farby","Betondur OUTDOOR powloka akrylowa na beton (zewnatrz)",
     "Scley Tasma malarska Washi 572, 48 mm x 33 m","14,78 zl",
     "1 - obowiazkowe",
     "Maskowanie krawedzi tarasu, balkonu, obrzuzy. Tasma Washi (do 60 dni) zachowuje wlasciwosci na zewnatrz przy przerwach pogodowych."),
    (17,"Farby","Betondur OUTDOOR powloka akrylowa na beton (zewnatrz)",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "2 - zalecane",
     "Krawedzie i narozniki na zewnatrz – wykončenia przy balustradach, scianach, stopniach."),
    (17,"Farby","Betondur OUTDOOR powloka akrylowa na beton (zewnatrz)",
     "Scley Tektura malarska falista 150 g/m2, 1x15 m","26,85 zl",
     "2 - zalecane",
     "Ochrona mebli ogrodowych, roslin, scian zewnetrznych. Tektura (15 m!) trwalsza od folii przy pracach zewnetrznych z wiatrem."),

    # ── 18. Impregnat WPC ─────────────────────────────────────────────────────
    (18,"Inne","Betondur impregnat do tarasow WPC kompozyt",
     "Hardy Walek malarski Hardstar 25 cm, runo 13 mm","17,81 zl",
     "1 - obowiazkowe",
     "Impregnat WPC naklada sie walkiem wzdluz desek. Runo 13 mm (grubsze) wchodzi w strukture rowkowanych desek kompozytowych zapewniajac rowne nasaczenie."),
    (18,"Inne","Betondur impregnat do tarasow WPC kompozyt",
     "Hardy Gabka glazurnicza 23x12x7 cm","24,49 zl",
     "1 - obowiazkowe",
     "Gabka strukturalna skutecznie wnika w rowki desek WPC docierajac do kazdego miejsca. Uzywana do wtarcia impregnatu w glebokie tekstury kompozytu."),
    (18,"Inne","Betondur impregnat do tarasow WPC kompozyt",
     "Hardy Kuweta malarska 26x35 cm (4\")","7,44 zl",
     "1 - obowiazkowe",
     "Do nakladania walkiem – niezbedna. Niska cena."),
    (18,"Inne","Betondur impregnat do tarasow WPC kompozyt",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "2 - zalecane",
     "Wykončenia przy balustradach, slupkach i krawedzziach tarasu gdzie walek nie dociera."),
    (18,"Inne","Betondur impregnat do tarasow WPC kompozyt",
     "Scley Tasma malarska Washi 572, 30 mm x 33 m","9,07 zl",
     "2 - zalecane",
     "Maskowanie balustrad, slupkow, scian – impregnatu nie powinno byc na elementach metalowych."),

    # ── 19. RDZA aktywator ────────────────────────────────────────────────────
    (19,"Inne","Betondur RDZA aktywator do efektu Rdzy (bezbarwny)",
     "Hardy Gabka glazurnicza 23x12x7 cm","24,49 zl",
     "1 - obowiazkowe",
     "Aktywator naklada sie lub spryskuje i wtiera gabka w nieregularny sposob – tym wlasnie uzyskuje sie efekt korozji. To glowne narzedzie do pracy z aktywatorem."),
    (19,"Inne","Betondur RDZA aktywator do efektu Rdzy (bezbarwny)",
     "Hardy Folia ochronna LDPE 4x5 m z regranulatu","8,63 zl",
     "1 - obowiazkowe",
     "Aktywator moze trwale zabarwic podloge. Ochrona obowiazkowa przy jego stosowaniu."),
    (19,"Inne","Betondur RDZA aktywator do efektu Rdzy (bezbarwny)",
     "Scley Tasma malarska Washi 572, 48 mm x 33 m","14,78 zl",
     "1 - obowiazkowe",
     "Aktywator jest agresywny chemicznie – tasma chroni sasiendne powierzchnie przed niezamierzonym zabarwieniem."),

    # ── 20. Renowator WPC ─────────────────────────────────────────────────────
    (20,"Inne","Betondur renowator do tarasow WPC kompozyt",
     "Hardy Walek malarski Hardstar 25 cm, runo 13 mm","17,81 zl",
     "1 - obowiazkowe",
     "Renowator naklada sie walkiem wzdluz desek – tak samo jak impregnat. Runo 13 mm wnika w structure WPC."),
    (20,"Inne","Betondur renowator do tarasow WPC kompozyt",
     "Hardy Gabka glazurnicza 23x12x7 cm","24,49 zl",
     "1 - obowiazkowe",
     "Do wtarcia renowatora w rowki desek WPC i wyrownywania koloru po naniesieniu."),
    (20,"Inne","Betondur renowator do tarasow WPC kompozyt",
     "Hardy Kuweta malarska 26x35 cm (4\")","7,44 zl",
     "1 - obowiazkowe",
     "Niezbedna do walka."),
    (20,"Inne","Betondur renowator do tarasow WPC kompozyt",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "2 - zalecane",
     "Wykončenia krawedzi i miejsc przy balustradach."),

    # ── 21. Srodek do plam WPC ────────────────────────────────────────────────
    (21,"Inne","Betondur srodek do plam WPC kompozyt",
     "Hardy Gabka glazurnicza 23x12x7 cm","24,49 zl",
     "1 - obowiazkowe",
     "Srodek do plam stosuje sie gabka – szorowanie plam z rowkowanych desek WPC. Gabka strukturalna dociera do glebszych warstw kompozytu, gdzie osadzaja sie plamy."),
    (21,"Inne","Betondur srodek do plam WPC kompozyt",
     "Scley Tasma malarska Washi 572, 30 mm x 33 m","9,07 zl",
     "2 - zalecane",
     "Ochrona balustrad i sasiednich elementow przy intensywnym szorowanie."),

    # ── 22. Srodek myjacy WPC ─────────────────────────────────────────────────
    (22,"Inne","Betondur srodek myjacy do tarasow WPC kompozyt",
     "Hardy Gabka glazurnicza 23x12x7 cm","24,49 zl",
     "1 - obowiazkowe",
     "Srodek myjacy WPC stosuje sie gabka lub szczotka. Gabka 23x12 cm to idealny rozmiar do mycia deski po desce."),
    (22,"Inne","Betondur srodek myjacy do tarasow WPC kompozyt",
     "Hardy Kuweta malarska 26x35 cm (4\")","7,44 zl",
     "2 - zalecane",
     "Miska do rozcienczania srodka myjacego i moczenia gabki. Praktyczne zastosowanie przy myciu tarasu."),

    # ── 23. Zywica epoksydowa ─────────────────────────────────────────────────
    (23,"Inne","BETONDUR Zywica Epoksydowa Bezbarwna + Utwardzacz 800g",
     "Hardy Szpachelka malarska 8 cm, stal nierdzewna, seria 78","12,51 zl",
     "1 - obowiazkowe",
     "Zywica epoksydowa nakladana szpachelka lub paca – standard aplikacji. Szpachelka ze stali nierdzewnej (odporna na chemikalia) to wlasciwe narzedzie. Bez szpachelki nie mozna rowniemiernie rozprowadzic zywicy."),
    (23,"Inne","BETONDUR Zywica Epoksydowa Bezbarwna + Utwardzacz 800g",
     "Hardy Pedzel plaski M7, 40 mm, seria 37","5,51 zl",
     "1 - obowiazkowe",
     "Zywica w naroznnikach i kraweddziach nakladana pedzelm. Pierwsza warstwa (bonding coat) czesto nakladana wlasnie pedzelm."),
    (23,"Inne","BETONDUR Zywica Epoksydowa Bezbarwna + Utwardzacz 800g",
     "Hardy Folia ochronna HDPE 4x5 m Standard","15,97 zl",
     "1 - obowiazkowe",
     "Zywica epoksydowa ekstremalnie trudna do usuniecia po utwardzeniu (7 dni!). HDPE (mocna folia) – konieczna. Zywica moze przebic cienksza folia LDPE."),
    (23,"Inne","BETONDUR Zywica Epoksydowa Bezbarwna + Utwardzacz 800g",
     "Scley Tasma malarska Washi 572, 48 mm x 33 m","14,78 zl",
     "1 - obowiazkowe",
     "Zywica na ograniczonych obszarach wymaga precyzyjnego maskowania. Szeroka 48 mm zapobiega wyciekaniu. Przy zywicy nie ma miejsca na bledy – utwardzonej nie da sie usunac bez szlifowania."),
    (23,"Inne","BETONDUR Zywica Epoksydowa Bezbarwna + Utwardzacz 800g",
     "Hardy Kuweta malarska 26x35 cm (4\")","7,44 zl",
     "2 - zalecane",
     "Do mieszania zywicy z utwardzaczem i nakladania. Kuweta staje sie jednorazowym naczyniemroboczym (zywica utwardza sie w pojemniku)."),
    (23,"Inne","BETONDUR Zywica Epoksydowa Bezbarwna + Utwardzacz 800g",
     "Hardy Wkladka do kuwety malarskiej 26x35 cm – przezroczysta","3,03 zl",
     "2 - zalecane",
     "Wkladka pozwala uzyc tej samej kuwety wielokrotnie – kluczowe przy pracy z zywica (brudna kuweta po pierwszej mieszance jest bezuzyteczna). Niska cena = latwy cross-sell."),
]

# ══════════════════════════════════════════════════════════════════════════════
# ARKUSZ 1 — PELNA TABELA
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Zestawy akcesoriow"
ws.freeze_panes = "A3"
ws.sheet_view.showGridLines = False

# ── tytuł ──
ws.merge_cells("A1:G1")
tc = ws["A1"]
tc.value = "BETONDUR.EU – Propozycje zestawów akcesoriów do produktów (wszystkie 23 produkty)"
tc.fill = fill(DARK_BLUE)
tc.font = Font(color="FFFFFF", bold=True, size=12)
tc.alignment = center()
ws.row_dimensions[1].height = 26

# ── nagłówki ──
hdrs = [
    ("Lp.", 5),
    ("Kategoria w sklepie", 20),
    ("Produkt główny (Betondur)", 38),
    ("Akcesorium do zestawu", 38),
    ("Cena akc.", 11),
    ("Priorytet", 18),
    ("Uzasadnienie", 70),
]
for ci, (h, w) in enumerate(hdrs, 1):
    c = ws.cell(row=2, column=ci, value=h)
    c.fill = fill(MID_BLUE)
    c.font = Font(color="FFFFFF", bold=True, size=9)
    c.alignment = center()
    c.border = border
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[2].height = 28

# ── wiersze danych ──
current_prod = None
group_idx = -1

for i, (nr, kat, prod, acc, cena, prio, uzas) in enumerate(rows, 1):
    r = i + 2
    if prod != current_prod:
        current_prod = prod
        group_idx += 1
    gi = group_idx % len(GROUP_COLORS)
    bg_even, bg_odd = GROUP_COLORS[gi]
    bg = bg_even if (i % 2 == 0) else bg_odd

    values = [nr, kat, prod, acc, cena, prio, uzas]
    for ci, v in enumerate(values, 1):
        c = ws.cell(row=r, column=ci, value=v)
        c.fill = fill(bg)
        c.border = border
        c.alignment = left() if ci in (2,3,4,7) else center()
        c.font = Font(size=9, bold=(ci == 3))
        if ci == 6:
            if "1" in v:
                c.fill = fill(GREEN_BG)
                c.font = Font(size=9, bold=True, color=GREEN_FG)
            else:
                c.fill = fill(YELLOW_BG)
                c.font = Font(size=9, color=YELLOW_FG)
    ws.row_dimensions[r].height = 50

# ══════════════════════════════════════════════════════════════════════════════
# ARKUSZ 2 — MATRYCA SKROCONA (produkty x akcesoria)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Matryca – widok skrócony")
ws2.freeze_panes = "B3"
ws2.sheet_view.showGridLines = False

# zbierz unikalowe produkty i akcesoria
from collections import OrderedDict
prod_list = list(OrderedDict.fromkeys(r[2] for r in rows))
acc_list  = list(OrderedDict.fromkeys(r[3] for r in rows))

# mapa: (prod, acc) -> priorytet
pmap = {}
for nr, kat, prod, acc, cena, prio, uzas in rows:
    pmap[(prod, acc)] = prio

ws2.merge_cells(f"A1:{get_column_letter(len(acc_list)+1)}1")
t2 = ws2["A1"]
t2.value = "BETONDUR.EU – Matryca priorytetów: produkty vs akcesoria"
t2.fill = fill(DARK_BLUE); t2.font = Font(color="FFFFFF", bold=True, size=11)
t2.alignment = center()
ws2.row_dimensions[1].height = 24

ws2.cell(row=2, column=1, value="Produkt Betondur").fill = fill(MID_BLUE)
ws2.cell(row=2, column=1).font = Font(color="FFFFFF", bold=True, size=8)
ws2.cell(row=2, column=1).alignment = center()
ws2.cell(row=2, column=1).border = border
ws2.column_dimensions["A"].width = 40

for ci, acc in enumerate(acc_list, 2):
    c = ws2.cell(row=2, column=ci, value=acc)
    c.fill = fill(MID_BLUE); c.font = Font(color="FFFFFF", bold=True, size=7)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=0)
    c.border = border
    ws2.column_dimensions[get_column_letter(ci)].width = 13
ws2.row_dimensions[2].height = 80

for ri, prod in enumerate(prod_list, 3):
    bg = "F8FBFF" if ri % 2 == 0 else "FFFFFF"
    c = ws2.cell(row=ri, column=1, value=prod)
    c.fill = fill(bg); c.font = Font(size=8, bold=True)
    c.alignment = left(); c.border = border
    ws2.row_dimensions[ri].height = 28
    for ci, acc in enumerate(acc_list, 2):
        val = pmap.get((prod, acc), "")
        cell = ws2.cell(row=ri, column=ci, value="")
        if "1" in val:
            cell.value = "●"
            cell.fill = fill(GREEN_BG); cell.font = Font(size=11, bold=True, color=GREEN_FG)
        elif "2" in val:
            cell.value = "○"
            cell.fill = fill(YELLOW_BG); cell.font = Font(size=11, color=YELLOW_FG)
        else:
            cell.fill = fill("F5F5F5"); cell.font = Font(size=9, color="CCCCCC")
            cell.value = "–"
        cell.alignment = center(); cell.border = border

# legenda
legend_row = len(prod_list) + 4
ws2.cell(row=legend_row, column=1, value="LEGENDA:").font = Font(bold=True, size=9)
ws2.cell(row=legend_row, column=2, value="●  Obowiązkowe (Priorytet 1)").fill = fill(GREEN_BG)
ws2.cell(row=legend_row, column=2).font = Font(size=9, bold=True, color=GREEN_FG)
ws2.cell(row=legend_row, column=3, value="○  Zalecane (Priorytet 2)").fill = fill(YELLOW_BG)
ws2.cell(row=legend_row, column=3).font = Font(size=9, color=YELLOW_FG)

# ══════════════════════════════════════════════════════════════════════════════
# ZAPIS
# ══════════════════════════════════════════════════════════════════════════════
out = "/Users/agencjaaqq/Desktop/Betondur_Zestawy_Akcesoriow_v2.xlsx"
wb.save(out)
print(f"OK: {out}")
